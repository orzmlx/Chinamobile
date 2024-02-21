import pandas as pd
import openpyxl
from openpyxl.chart.series import Series
import numpy as np
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook
import math
import huaweiutils
import os
import logging


class reporter:

    def __init__(self, check_result, outpath, standard_path, cities, date):
        self.date = date
        self.result = check_result
        self.outpath = outpath
        self.cities = cities
        self.standard_path = standard_path
        self.manufacturers = ['huawei', 'zte', 'ericsson']
        self.point_standard_df = pd.read_excel(self.standard_path, sheet_name='小区级别核查配置', true_values=["是"],
                                               false_values=["否"], dtype=str)
        self.multi_standard_df = pd.read_excel(self.standard_path, sheet_name='频点级别核查配置', true_values=["是"],
                                               false_values=["否"], dtype=str)
        # self.writer = pd.ExcelWriter(outpath, engine='xlsxwriter')  # 创建pandas.ExcelWriter实例，赋值给writer
        self.params = self.point_standard_df['参数名称'].unique().tolist()
        self.params.extend(self.multi_standard_df['参数名称'].unique().tolist())
        self.g5_statistic_dict = {}
        self.g4_statistic_dict = {}
        pt_params = self.point_standard_df[["类别", "参数名称", "原始参数名称"]]
        freq_param = self.multi_standard_df[["类别", "参数名称", "原始参数名称"]]
        params_df = pd.concat([pt_params, freq_param], axis=0)
        params_df.dropna(how='all', inplace=True, axis=0)
        self.params_df = params_df

    def create_result_header(self, sheet, start_col, row):
        # 先生成层级
        clzz = np.array(self.params_df['类别'].unique().tolist())
        clzz = [x for x in clzz if not x == 'nan']
        # start_col = 4
        class_colors = ['48D1CC', '00FA9A', 'FFA500', '1E90FF', '800080']
        for i in range(len(clzz)):
            # 该层级下有那些参数
            c = clzz[i]
            j = i
            if i >= len(class_colors):
                j = 0
            color = class_colors[j]
            param_names = self.params_df[self.params_df['类别'] == c]['参数名称'].unique().tolist()
            self.create_business_sector(c, sheet, start_col, param_names, row, color)
            start_col = start_col + len(param_names)
        return start_col - 1

    def create_business_sector(self, c, sheet, start_col, param_names, row, color):
        # 先把参数写上去
        param_start_col = start_col
        param_row = row + 1
        for p in param_names:
            cell = sheet.cell(row=param_row, column=param_start_col, value=p)
            cell.font = Font(u'微软雅黑', size=9, bold=True, color='00000000')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            param_start_col = param_start_col + 1
        # 在写层级上去
        end_columns = start_col + len(param_names) - 1
        sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_columns)
        cell = sheet.cell(row=row, column=start_col, value=c)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(u'微软雅黑', size=10, bold=True, color='00000000')
        cell.fill = PatternFill('solid', fgColor=color)

    def output_general_check_result(self):
        self.get_all_statistic()
        # if not os.path.exists(self.report_path):
        # 创建一个工作簿
        wb = openpyxl.Workbook()
        # wb = openpyxl.load_workbook(filename=self.standard_path)
        # sheet = wb.get_sheet_by_name('核查结果_test')
        sheet = wb.active
        # all_params = self.param.dif['参数名称'].unique().tolist()
        start_row = 1
        start_col = 1
        common_cell_list = [['地市', '不合规总数'], ['地市', '核查总数'], ['地市', '总合规率'],
                            ['地市', '不合规总数'], ['地市', '核查项总数'], ['地市', '总合规率']]
        headers = ['4G小区不合规数量', '4G小区核查项总数', '4G小区参数配置核查合规率',
                   '5G小区不合规数量', '5G小区核查项总数', '5G小区参数配置核查合规率']
        for index, header_value in enumerate(headers):
            sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row,
                              end_column=len(common_cell_list[index]) + len(self.params) - 1)
            cell = sheet.cell(row=start_row, column=start_col, value=header_value)
            cell.font = Font(u'微软雅黑', size=10, bold=True, color='00000000')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill('solid', fgColor='FFFF00')
            self.create_common_cell(common_cell_list[index], sheet, 1, start_row + 1)
            end_col = self.create_result_header(sheet, len(common_cell_list[index]) + 1, start_row + 1)
            city_row = start_row + 3
            for c in self.cities:
                cell = sheet.cell(row=city_row, column=1, value=c)
                self.write_city_statistic_data(sheet, c, city_row, end_col, header_value)
                cell.font = Font(u'微软雅黑', size=10, bold=True, color='00000000')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                city_row = city_row + 1
            start_row = start_row + 2 + len(self.cities) + 1
            summary_row = city_row - 1
            self.write_summary_statistic(sheet, summary_row, end_col, header_value)
        report_outpath = os.path.join(self.outpath, '互操作参数核查结果' + self.date + '.xlsx')
        wb.save(report_outpath)

    def write_summary_statistic(self, sheet, summary_row, end_col, header_value):
        statistic_dict = self.get_city_dict(header_value, None)
        if len(statistic_dict) == 0:
            return
        param_summary_dict = self.get_param_summary_dict(statistic_dict)
        summary_value = 0
        for i in range(end_col):
            if i == 0 or i == 1 or i == 2:
                continue
            param_name = sheet.cell(row=summary_row - len(self.cities), column=i).value
            param_stat = param_summary_dict[param_name]
            if param_stat[0] == 0 and param_stat[1] == 0:
                continue
            if header_value.find('不合规数量') >= 0:
                write_value = param_stat[1]
                sheet.cell(row=summary_row, column=i, value=write_value)
            elif header_value.find('核查项总数') >= 0:
                write_value = param_stat[1] + param_stat[0]
                sheet.cell(row=summary_row, column=i, value=write_value)
            elif header_value.find('合规率') >= 0:
                if param_stat[1] + param_stat[0] == 0:
                    write_value = 0
                else:
                    write_value = param_stat[0] / (param_stat[1] + param_stat[0])
                sheet.cell(row=summary_row, column=i, value=str("{:.2%}".format(write_value)))
            else:
                raise Exception(header_value + '没有合适的数据汇集方法')
            if header_value.find('合规率') >= 0:
                summary_value = param_summary_dict['summary'][0] / (
                        param_summary_dict['summary'][1] + param_summary_dict['summary'][0])
                summary_value = str("{:.2%}".format(summary_value))
            else:
                summary_value = summary_value + int(write_value)
        sheet.cell(row=summary_row, column=2, value=summary_value)

    def get_param_summary_dict(self, statistic_dict):
        param_summary_dict = {}
        all_param_qualified_sum = 0
        all_param_unqualified_sum = 0
        for p in self.params:
            param_qualified_sum = 0
            param_unqualified_sum = 0
            for city in statistic_dict:
                city_statistic = statistic_dict[city]
                if p not in city_statistic:
                    logging.debug("统计中没有发现参数:" + p)
                    continue
                param_statistic = city_statistic[p]
                qualified_value = param_statistic[0]
                unqualified_value = param_statistic[1]
                param_qualified_sum = param_qualified_sum + qualified_value
                param_unqualified_sum = param_unqualified_sum + unqualified_value
            param_summary_dict[p] = (param_qualified_sum, param_unqualified_sum)
            all_param_qualified_sum = all_param_qualified_sum + param_qualified_sum
            all_param_unqualified_sum = all_param_unqualified_sum + param_unqualified_sum
        param_summary_dict['summary'] = (all_param_qualified_sum, all_param_unqualified_sum)
        return param_summary_dict

    def get_all_statistic(self):
        g5_freq_stat_dict = self.statistic_city_data('5G', '_freq.csv')
        g5_cell_stat_dict = self.statistic_city_data('5G', '_cell.csv')
        g4_freq_stat_dict = self.statistic_city_data('4G', '_freq.csv')
        g4_cell_stat_dict = self.statistic_city_data('4G', '_cell.csv')
        self.g5_statistic_dict = self.merge_city_statistic(g5_freq_stat_dict, g5_cell_stat_dict)
        self.g4_statistic_dict = self.merge_city_statistic(g4_freq_stat_dict, g4_cell_stat_dict)

    def merge_city_statistic(self, g5_freq_stat_dict, g5_cell_stat_dict):
        freq_cities = list(g5_freq_stat_dict.keys())
        cell_cities = list(g5_cell_stat_dict.keys())
        freq_cities.extend(cell_cities)
        all_unique_cities = set(freq_cities)
        res_dict = {}
        for city in all_unique_cities:
            cell_param_dict = g5_cell_stat_dict.get(city, {})
            freq_param_dict = g5_freq_stat_dict.get(city, {})
            param_dict = {**cell_param_dict, **freq_param_dict}
            res_dict[city] = param_dict
        return res_dict

    def statistic_city_data(self, system, suffix):
        all_flist = []
        for m in self.manufacturers:
            find_path = os.path.join(self.outpath, m, self.date, system)
            flist = huaweiutils.find_file(find_path, suffix)
            all_flist.extend(flist)
        city_res_dict = {}
        if len(all_flist) == 0:
            logging.info(self.outpath + "下没有找到包含" + suffix + '的文件')
            return city_res_dict
        for f in all_flist:
            df = pd.read_csv(f)
            self.statistic_data(df, city_res_dict)
        return city_res_dict

    def create_common_cell(self, common_cell, sheet, start_col, row):
        # colors = ['#000080', '#B22222', '#2E8B57']

        for index, c in enumerate(common_cell):
            sheet.merge_cells(start_row=row, start_column=start_col, end_row=row + 1, end_column=start_col)
            cell = sheet.cell(row=row, column=start_col, value=c)
            cell.font = Font(u'微软雅黑', size=9, bold=True, color='00000000')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            start_col = start_col + 1

    def statistic_data(self, df, city_res_dict):
        """
            统计小区级别参数的合规率和检查总数
            return 按地市返回合格和不合格的总数
        """
        cols = df.columns.tolist()
        judge_cols = [c for c in cols if c.find('合规') >= 0]
        judge_cols.append('地市')
        df = df[judge_cols]
        grouped = df.groupby('地市')
        for city, g in grouped:
            city_res = city_res_dict.get(city, {})
            for c in judge_cols:
                if c == '地市':
                    continue
                g[c] = g[c]
                c_qualified_num = len(g[g[c] == True])
                c_unqualidied_num = len(g[g[c] == False])
                c = c.split('#')[0]
                # 将简化名称放入字典
                c = self.params_df[self.params_df['原始参数名称'] == c]['参数名称'].iloc[0]
                param_tuple = city_res.get(c, (0, 0))
                n_qualified_num = param_tuple[0] + c_qualified_num
                n_unqualidied_num = param_tuple[1] + c_unqualidied_num
                city_res[c] = (n_qualified_num, n_unqualidied_num)
            city_res_dict[city] = city_res
        return city_res_dict

    def find_statistic(self, city_dict, param, header_value):
        statistic_tuple = city_dict.get(param, None)
        # 如果两个数据都是0，那么肯定在另外一个freq字典中
        if statistic_tuple is None:
            return 0, 0

        unqualified_num = statistic_tuple[1]
        qualified_num = statistic_tuple[0]
        return qualified_num, unqualified_num

    def get_city_dict(self, header_value, city):
        if header_value.find('4G') >= 0:
            city_dict = self.g4_statistic_dict.get(city, {}) if city is not None else self.g4_statistic_dict
        elif header_value.find('5G') >= 0:
            city_dict = self.g5_statistic_dict.get(city, {}) if city is not None else self.g5_statistic_dict
        else:
            raise Exception(header_value + "找不到对应的统计数据")
        return city_dict

    def write_city_statistic_data(self, sheet, city, city_row, end_col, header_value):
        # if header_value.find('4G') >= 0:
        #     city_dict = self.g4_statistic_dict.get(city, {})
        # elif header_value.find('5G') >= 0:
        #     city_dict = self.g5_statistic_dict.get(city, {})
        # else:
        #     raise Exception(header_value + "找不到对应的统计数据")
        city_dict = self.get_city_dict(header_value, city)
        if 0 == len(city_dict):
            logging.info("没有找到【" + city + "】的统计数据,请检查统计表")
            return
        unqualified_sum = 0
        qualified_sum = 0
        for i in range(end_col):
            if i == 0:
                continue
            # 判断属性名称
            cell = sheet.cell(row=3, column=i)
            param = cell.value
            if param not in city_dict.keys():
                continue
            statistic_tuple = city_dict[param]
            unqualified_num = statistic_tuple[1]
            qualified_num = statistic_tuple[0]
            if qualified_num == 0 and unqualified_num == 0:
                if statistic_tuple[0] == 0 and statistic_tuple[1] == 0:
                    raise Exception("没有找到参数【" + param + "】在城市【" + city + "】的统计数据,数据均为默认值0")
            unqualified_sum = unqualified_sum + unqualified_num
            qualified_sum = qualified_sum + qualified_num
            if header_value == '4G小区不合规数量':
                sheet.cell(row=city_row, column=i, value=unqualified_num)
            elif header_value == '4G小区核查项总数':
                sheet.cell(row=city_row, column=i, value=unqualified_num + qualified_num)
            elif header_value == '4G小区参数配置核查合规率':
                sheet.cell(row=city_row, column=i,
                           value=str((qualified_num / (unqualified_num + qualified_num))))
            elif header_value == '5G小区核查项总数':
                sheet.cell(row=city_row, column=i, value=unqualified_num + qualified_num)
            elif header_value == '5G小区参数配置核查合规率':
                sheet.cell(row=city_row, column=i,
                           value=str("{:.2%}".format((qualified_num / (unqualified_num + qualified_num)))))
            elif header_value == '5G小区不合规数量':
                sheet.cell(row=city_row, column=i, value=unqualified_num)
        if header_value.find('不合规数量') >= 0:
            sheet.cell(row=city_row, column=2, value=unqualified_sum)
        elif header_value.find('核查项总数') >= 0:
            sheet.cell(row=city_row, column=2, value=unqualified_sum + qualified_sum)
        elif header_value.find('合规率') >= 0:
            sheet.cell(row=city_row, column=2,
                       value=str("{:.2%}".format((qualified_sum / (qualified_sum + unqualified_sum)))))


if __name__ == "__main__":
    check_result = 'C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\huawei\\result\\check_result.csv'
    # 所有的数据，包括中兴，华为，爱立信
    outpath = 'C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\huawei\\result'
    standard_path = "C:\\Users\\No.1\\Desktop\\teleccom\\互操作参数核查结果.xlsx"
    cities = ['湖州', '杭州', '金华', '嘉兴', '丽水', '宁波', '衢州', '绍兴', '台州', '温州', '舟山', '汇总']
    report_path = "C:\\Users\\No.1\\Desktop\\teleccom\\互操作参数核查结果1.xlsx"
    date = '20240121'
    reporter = reporter(check_result, outpath, standard_path, cities, date)
    reporter.output_general_check_result()
# cell_stat_res = reporter.statistic_cell('5G')
# cell_stat_res = reporter.statistic_freq('5G')
