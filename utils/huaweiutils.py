# -*- coding:utf-8 -*-
import math
import copy
import itertools
import logging
import os
import pathlib
import shutil
from copy import deepcopy
from difflib import SequenceMatcher

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame
from tqdm import tqdm

from configuration import zte_configuration, huawei_configuration

replace_char = ['秒', 'dB']

logging.basicConfig(format='%(asctime)s : %(message)s', datefmt='%m/%d/%Y %I:%M:%S', level=logging.INFO)


def split_csv(path, chunksize):
    files = find_file(path, '.csv')
    for index, f in enumerate(files):
        dfs = pd.read_csv(f, chunksize=chunksize, encoding='gbk')
        f_name = f.name
        merge_result = pd.DataFrame()
        i = 0
        for index0, chunk in enumerate(dfs):
            merge_result = chunk if merge_result.empty else pd.concat([merge_result, chunk], axis=0)
            if i >= 3:
                merge_result.to_csv(os.path.join(path, f_name + '_' + str(index0) + '.csv'), index=False,
                                    encoding='utf_8_sig')
                merge_result = pd.DataFrame()
                i = -1
            i = i + 1
        merge_result.to_csv(os.path.join(path, f_name + '_last' + '.csv'), index=False, encoding='utf_8_sig')


def combine_file_by_name(path):
    f_dict = {}
    all_file = find_file(path, '.csv')
    for i in range(len(all_file)):
        if 'temp' == os.path.split(os.path.abspath(os.path.dirname(all_file[i])))[1]:
            continue
        f_name = all_file[i].name
        f_list = f_dict.get(f_name, [])
        f_list.append(str(all_file[i]))
        f_dict[f_name] = f_list
    for key in f_dict:
        merge_res = pd.DataFrame()
        f_list = f_dict[key]
        for f in f_list:
            tqdm.pandas(desc='Reading CSV')
            # 读取CSV文件，tqdm将自动处理进度条显示
            df = pd.read_csv(f)
            merge_res = df if merge_res.empty else pd.concat([merge_res, df], axis=0)
        output_path = os.path.join(path, "temp_combine")
        if not os.path.exists(output_path):
            os.makedirs(output_path)
        merge_res.to_csv(os.path.join(output_path, str(key)), index=False, encoding='utf_8_sig')


def create_header(df, path, class_dict, base_cols):
    # wb = load_workbook(path)
    # sheet = wb.active
    logging.info(">>>>>修改结果表头....<<<<<")
    wb = Workbook()
    sheet = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)
    # 从第一行开始,插入两行
    sheet.insert_rows(1, 2)
    base_col_len = len(base_cols)
    first_start_row = 1
    second_start_row = 2
    header_row = 3
    first_start_col = base_col_len + 1
    second_start_col = first_start_col
    first_class_colors = ['48D1CC', '00FA9A', 'FFA500', '1E90FF', '800080']
    second_class_colors = ['7B68EE', '6495ED', '48D1CC', '98FB98', 'F0E68C']
    header_colors = ['DCDCDC', 'FAEBD7', 'FFFFE0', 'D4F2E7', 'E6E6FA']
    second_color_index = 0
    for i, clzz in enumerate(class_dict.keys()):
        # clzz_cols = class_dict[clzz]
        second_dict = class_dict[clzz]
        first_clzz_values = second_dict.values()
        first_clzz_cols = list(itertools.chain(*first_clzz_values))
        first_end_column = first_start_col + len(first_clzz_cols) - 1
        first_end_row = first_start_row
        m = i
        if i >= len(first_class_colors):
            m = 0
        first_color = first_class_colors[m]
        sheet.merge_cells(start_row=first_start_row, start_column=first_start_col, end_row=first_end_row,
                          end_column=first_end_column)
        cell = sheet.cell(row=first_start_row, column=first_start_col, value=clzz)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(u'微软雅黑', size=10, bold=True, color='00000000')
        cell.fill = PatternFill('solid', fgColor=first_color)
        first_start_col = first_end_column + 1
        for j, second_clzz in enumerate(second_dict.keys()):
            second_end_row = second_start_row
            second_clzz_cols = len(second_dict[second_clzz])
            second_end_column = second_start_col + second_clzz_cols - 1
            if second_color_index >= len(second_class_colors) - 1:
                second_color_index = 0
            second_color = second_class_colors[second_color_index]
            sheet.merge_cells(start_row=second_start_row, start_column=second_start_col, end_row=second_end_row,
                              end_column=second_end_column)
            cell = sheet.cell(row=second_start_row, column=second_start_col, value=second_clzz)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(u'微软雅黑', size=10, bold=True, color='00000000')
            cell.fill = PatternFill('solid', fgColor=second_color)
            for row in sheet.iter_rows(min_row=header_row, min_col=second_start_col, max_col=second_end_column,
                                       max_row=header_row):
                for header_cell in row:
                    header_cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='left')
                    header_cell.fill = PatternFill('solid', fgColor=header_colors[second_color_index])
                    header_cell.font = Font(u'微软雅黑', size=9, bold=True, color='00000000')
            second_color_index = second_color_index + 1
            second_start_col = second_end_column + 1
    # 获取第三行
    header_row = sheet[header_row]
    for index, cell in enumerate(header_row):
        cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='left')
        cell.fill = PatternFill('solid', fgColor='E1FFFF')
        cell.font = Font(u'微软雅黑', size=9, bold=True, color='00000000')
        if index == len(base_cols) - 1:
            break
    wb.save(os.path.join(os.path.split(path)[0], '互操作小区级核查结果.csv'))


def unzip_all_files(path, dest_path=None, zipped_file=[]):
    """
     解压原始log文件到目标文件夹，防止解压文件中包含解压文件，进行递归解压，当解压文件数量没有提升
    :param path:
    :param dest_path:
    :param zipped_file:
    :return:
    """
    logging.info('==============开始解压路径' + str(path) + '下文件==============')
    if path == '' or path is None:
        return
    os.chmod(path, 7)
    if dest_path is not None:
        os.chmod(dest_path, 7)
    zip_files = find_file(path, '.zip')
    if 0 < len(zipped_file) == len(zip_files):
        return
    if len(zip_files) == 0:
        return
    for file in zip_files:
        base_name = os.path.basename(file)
        if base_name in zipped_file:
            continue
        dest_dir = str(file).replace('.zip', '') \
            if dest_path is None else os.path.join(dest_path, base_name.replace('.zip', ''))

        # patoolib.extract_archive(file,dest_dir)
        try:
            shutil.unpack_archive(file, extract_dir=dest_dir)
            # os.remove(file)
        except Exception as e:
            logging.info(e)
        zipped_file.append(base_name)
    unzip_all_files(path, dest_path, zipped_file)

    # f = zipfile.ZipFile(f, 'r')  # 压缩文件位置


def get_content_col(base_cols, cols):
    diff = set(cols) - set(base_cols)
    if "对端频带" in diff:
        diff.remove("对端频带")
    diff1 = copy.deepcopy(diff)
    content_cols = []
    for m in diff:
        if m.find('#') < 0:
            content_cols.append(m)
            for n in diff1:
                if m == n:
                    continue
                if n.find(m) >= 0 and n.find('#') >= 0:
                    content_cols.append(n)
    return content_cols


def is_lists_of_same_length(dict_obj):
    """
    判断字典对应的列表是否都是相同长度，如果都相同，那么返回这个长度
    :param dict_obj:
    :return: int
    """
    if len(dict_obj) == 0:
        return 0
    first_key = next(iter(dict_obj))
    first_length = len(dict_obj[first_key])
    if all(len(dict_obj[key]) == first_length for key in dict_obj):
        return first_length
    else:
        return -1


def find_file(directory, file_extension):
    files = []
    if isinstance(directory, list):
        for d in directory:
            res = _find_file(d, file_extension)
            files.extend(res)
        return files
    elif isinstance(directory, str) or isinstance(directory, pathlib.Path):
        return _find_file(directory, file_extension)
    else:
        raise Exception(directory + '数据类型错误，无法遍历下面的文件')


def _find_file(directory, file_extension):
    files = []
    if os.path.isfile(directory):
        return []
    for item in pathlib.Path(directory).rglob('*'):
        if str(item).endswith(file_extension):
            files.append(item)
    return files


def output_csv(df, file_name, out_path, is_format):
    if not os.path.exists(out_path):
        os.makedirs(out_path)
    if is_format:
        file_name = remove_digit(file_name, ['=', ":"])
    file_out_path = os.path.join(out_path, file_name)
    if file_name in os.listdir(out_path):
        df.to_csv(file_out_path, index=False)
    else:
        df.to_csv(file_out_path, index=False)


def only_has_digtal_diff(str1, str2):
    """
        判断两个字符串不相同的部分
    """
    if str1 == str2:
        return False
    matcher = SequenceMatcher(None, str1, str2)
    match_codes = matcher.get_opcodes()
    replace_num = 0
    for tag, i, j, m, n in match_codes:
        if tag == 'delete' or tag == 'insert':
            return False
        if tag == 'replace':
            replace_num = replace_num + 1
            if not str1[i].isdigit() or not str2[m].isdigit():
                return False
    if replace_num > 1:
        return False
    return True


def add_4g_cgi(cell_df, enode_df):
    cell_df = cell_df[cell_df['NB-IoT小区指示'] != '是']
    cell_df = pd.merge(cell_df, enode_df[['网元', 'eNodeB标识']], on='网元')
    cell_df['CGI'] = "460-00-" + cell_df["eNodeB标识"].apply(str) + "-" + cell_df['小区标识'].apply(str)
    return cell_df


def add_5g_cgi(ducell_df, gnodeb_df):
    ducell_df = pd.merge(ducell_df, gnodeb_df[['网元', 'gNodeB标识']], on='网元')
    ducell_df.rename(columns={'NRDU小区标识': 'NR小区标识'}, inplace=True)

    # print(ducell_df.shape)
    return ducell_df


def remove_digit(str, add_characters):
    """
        去除数字和其它
    """
    result = ""
    if add_characters is None:
        add_characters = []
    for c in str:
        if c.isdigit() or c in add_characters:
            continue
        result += c
    return result


def remove_name(x):
    try:
        if str(x).find(":") < 0:
            return x
        values = str(x).split(':')
        name = values[1]
        return name
    except Exception as e:
        logging.error(e)
        return x


def flatten_features(df, c):
    # 取所有数据的第一行，看是否有子列
    # cols = df.columns.tolist()
    df1 = deepcopy(df)
    values = str(df1[c].iloc[0])
    if values.find("&") >= 0 or (values.find(":") >= 0 and values.find("开关") >= 0):
        edit_df = df1[c]
        df1.drop(c, axis=1, inplace=True)
        if values.find("&") >= 0:
            edit_df = edit_df.astype('str').str.split('&', expand=True)
            current_cols = edit_df.columns.tolist()
        else:
            edit_df = pd.DataFrame(edit_df)
            edit_df.rename(columns={c: 0}, inplace=True)
            current_cols = [0]
        first_row = edit_df.reset_index(drop=True).iloc[0]
        new_cols = []
        for col in current_cols:
            cell = first_row[col]
            # 如果有部分为空,那么需要不停往下检查，直到找到非空
            if pd.isna(cell):
                all_values = edit_df[col].tolist()
                for v in all_values:
                    if v is not None:
                        cell = v
                        break
            if cell.find(":") < 0:
                continue
            col_name = cell.split(":")[0]
            new_cols.append(col_name)
            edit_df[col] = edit_df[col].apply(remove_name)
        edit_df.reset_index(drop=True, inplace=True)
        edit_df.columns = new_cols
        df1 = pd.concat([df1, edit_df], axis=1)
    return df1


def parse_strategy_file(path):
    g5_data_strategy = pd.read_excel(path, sheet_name="自身推荐")
    g45_data_strategy = pd.read_excel(path, sheet_name="45G数据")
    g5_data_strategy = g5_data_strategy[g5_data_strategy['事件'] == 'A5']
    g5_data_strategy.drop('事件', axis=1, inplace=True)
    cols = g5_data_strategy.columns.tolist()
    cols0 = g45_data_strategy.columns.tolist()
    keep_name_list = ['频带', '厂家', '覆盖类型']
    for c in cols:
        if c in keep_name_list:
            continue
        g5_data_strategy.rename(columns={c: c + "#合规"}, inplace=True)
    for c in cols0:
        if c in keep_name_list:
            continue
        g45_data_strategy.rename(columns={c: c + "#合规"}, inplace=True)
    return g5_data_strategy, g45_data_strategy


def add_strategy_info(g5_data_strategy_df, g45_data_strategy_df, report_df):
    g5_cols = g5_data_strategy_df.columns.tolist()
    g45_cols = g45_data_strategy_df.columns.tolist()
    sum_cols = g5_cols + g45_cols
    unused_cols = ['厂家', '覆盖类型', '频带']
    feature_cols = [x.strip() for x in sum_cols if x not in unused_cols]
    original_feature_cols = [x.split("#")[0].strip() for x in feature_cols]
    report_cols = report_df.columns.tolist()
    report = report_df.merge(g5_data_strategy_df, how='left', on=['厂家', '覆盖类型', '频带'])
    report = report.merge(g45_data_strategy_df, how='left', on=['厂家', '频带'])
    # 对列重排序
    base_cols = ['网元', 'NR小区标识', 'NRDU小区名称', 'CGI', '频带', '双工模式', '覆盖类型', '覆盖场景', '厂家']
    diff_cols = list(set(report_cols).difference(set(original_feature_cols)))
    rest_cols = set(diff_cols) - set(base_cols)
    base_cols.extend(rest_cols)
    for index, c in enumerate(feature_cols):
        original_name = c.split("#")[0]
        # if report_cols.find(original_name) >= 0:
        if original_name in report_cols:
            new_col = '是否合规' + "_" + str(index)
            report[new_col] = report.apply(add_judgement, axis=1, args=(original_name, c))
            diff_cols.append(original_name)
            diff_cols.append(c)
            diff_cols.append(new_col)
    ##重新对列排序
    return report[diff_cols]


def judge(df, param):
    judge_res = []
    df.rename(columns={"推荐值": param + "#推荐值"}, inplace=True)
    for recommand, value in zip(df[param + "#推荐值"], df[param]):
        value = str(value)
        recommand = str(recommand)
        if value == 'nan':
            # judge_res.append("没有找到参考值")
            judge_res.append(True)
            continue
        if recommand == 'nan':
            #华为要求没有值不判断
            judge_res.append(math.nan)
            continue
        if value.find(';') >= 0:
            values = value.split(';')
            match = True
            for v in values:
                judge = get_judge(recommand, v)
                if not judge:
                    match = False
                    break
            judge_res.append(match)
        else:
            judge = get_judge(recommand, value)
            judge_res.append(judge)

    return judge_res


def get_judge(recommend, value):
    if 'nan' == recommend:
        return math.nan
    if recommend.find('[') >= 0 and recommend.find(']') >= 0:
        return range_judge(value, recommend)
    elif recommend.find(',') >= 0:
        return list_judge(value, recommend)
    else:
        return single_value_judge(value, recommend)


def linear_calculation(x, m1, b, m2):
    return x * m1 + b * m2


def mapToBand(x, band_dict):
    if 'nan' == str(x):
        return '其他频段'
    for key, item in band_dict.items():
        if str(x) in item:  # 证明该频段是4G频段
            return key
    return '其他频段'


def generate_4g_frequency_band_dict(df):
    # df.dropna(axis=0, inplace=True, how='any')
    band_list = []
    g4_freq_band_dict = {}
    for band, offset in zip(df['工作频段'], df['频率偏置'], df['中心载频信道号']):
        band = str(band)
        if pd.isna(band):
            continue
        band = band.replace('频段', '')
        if band.find('FDD') >= 0:
            band = str(offset)
            # 如果不是FDD-1800或者FDD-900,那么直接去掉数字
            if str(offset).find('FDD') < 0:
                band = remove_digit(band, [])
            if str(offset).find('-') >= 0:
                band = band.replace('-', '')
        band_list.append(band)
    return zte_configuration.g4_band_dict, band_list


def read_csv(file_name, usecols=None, dtype=None, manufacturer=None) -> DataFrame:
    try:
        res_df = pd.read_csv(file_name, usecols=usecols, dtype=dtype) if dtype is not None else pd.read_csv(file_name,
                                                                                                            usecols=usecols)
    except Exception as e:
        logging.info('读取文件' + file_name + '报错,读取的列名:' + str(usecols) + '报错信息:' + str(e))
        logging.info("读取文件失败,改用gbk编码方式重读csv文件")
        res_df = pd.read_csv(file_name, usecols=usecols, dtype=dtype, encoding='gbk') \
            if dtype is not None else pd.read_csv(file_name, usecols=usecols, encoding='gbk')
    # 中兴的header有的文件有两行，这里去掉一行
    if manufacturer == '中兴':
        skipheader = res_df.iloc[0]
        if 'CGI' in res_df.columns.tolist() and skipheader['CGI'] == 'CGI':
            res_df.drop(0, inplace=True)
    return res_df


def list_to_str(check_list):
    check_set = list(set(check_list))
    check_set = [c for c in check_set if str(c) != 'nan']
    result = '|'.join(check_set)
    if 'nan' not in check_list:
        result = result + '|nan'
    return result


def merge_dfs(lst, on, cell_identity):
    if len(lst) == 0:
        return pd.DataFrame()
    init_df = lst[0]
    for i in range(len(lst)):
        copy_on = copy.deepcopy(on)
        if i == 0:
            continue
        if cell_identity in lst[i].columns.tolist() and cell_identity in init_df:
            copy_on.append(cell_identity)
        init_df = init_df.merge(lst[i], how='left', on=copy_on)
    return init_df

def is_float(str):
    try:
        float(str)
        return True
    except ValueError:
        return False

def single_value_judge(x, standard):
    try:
        for c in replace_char:
            x = str(x).replace(c, "")
        if is_float(standard):
            standard = str(int(float(standard)))
        return str(x) == standard
    except Exception as e:
        logging.error(e)


def list_judge(x, standard):
    for c in replace_char:
        x = str(x).replace(c, "")
    splits = standard.split(',')
    return True if str(x) in splits else False


def range_judge(x, standard):
    for c in replace_char:
        x = str(x).replace(c, "")
    standard = standard.replace("[", "").replace("]", "")
    splits = standard.split(",")
    if len(splits) != 2:
        raise Exception("范围推荐值:【" + standard + "】格式不符合要求")
    else:
        try:
            min_value = int(splits[0])
            max_value = int(splits[1])
            if min_value > max_value:
                temp = max_value
                max_value = min_value
                min_value = temp
            if min_value <= int(float(x)) <= max_value:
                return True
            else:
                return False
        except Exception as e:
            raise Exception("范围推荐值:【" + standard + "】无法转换为数字", e)


def add_judgement(x, original_name, c):
    judgement_value = str(x[c])
    original_value = x[original_name]
    if len(judgement_value.strip()) == 0:
        return ""
    if judgement_value.find("~") >= 0:
        splits = judgement_value.split("~")
        value_0 = int(splits[0])
        value_1 = int(splits[1])
        temp = value_1
        if value_0 > value_1:
            value_1 = value_0
            value_0 = temp
        if value_0 < original_value < value_1:
            return True
        elif value_0 == judgement_value or value_1 == judgement_value:
            return judgement_value, True
        else:
            return False
    elif judgement_value == 'nan':
        return True
    elif judgement_value.find("[") >= 0 and judgement_value.find("]") >= 0:
        judgement_value.replace("[", "").replace("]", "")
    else:
        return True if int(float(judgement_value)) == int(
            float(original_value)) else False


if __name__ == "__main__":
    # report_path = "C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\huawei\\result\\all_result.csv"
    # report_df = pd.read_csv(report_path)
    # report_df['频带'] = report_df['频带'].map({"n41": "2.6G", "n28": "700M", "n78": "4.9G", "n79": "4.9G"})
    # strategy_path = "C:\\Users\\No.1\\Desktop\\teleccom\\互操作参数策略.xlsx"
    # g5_data_strategy, g45_data_strategy = parse_strategy_file(strategy_path)
    # report = add_strategy_info(g5_data_strategy, g45_data_strategy, report_df)
    # report_cols = report.columns.tolist()
    # report.to_csv("C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\huawei\\result\\report_strategy.csv",
    #               index=False, encoding='utf_8_sig')
    # str1 = 'LST NRCELLHOEUTRANMEAGRP:INTERRHOTOEUTRANMEASGRPID=2'
    # # str2 = 'LST NRCELLINTRAFHOMEAGRPINTRAFREQHOMEASGROUPID=2.csv'
    # str2 = 'LST NRCELLQCIBEARER:QCI=5'
    # # print(only_has_digtal_diff(str1, str2))
    # print(remove_digit(str1, ['=']))
    path = 'C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\huawei\\20240426\\5G'

    combine_file_by_name(path)
    # path = 'C:\\Users\\No.1\\Desktop\\分表'
    # split_csv(path, 200000)
    # path = 'C:\\Users\\No.1\\Desktop\\zip_test'
    # unzip_all_files(path)
    # huawei_txts = find_file(path, '.txt')
    # dest = 'C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\huawei\\20240327\\5G\\raw_data'
    # if not os.path.exists(dest):
    #     os.makedirs(dest)
    # for txt in huawei_txts:
    #     shutil.move(str(txt), dest)
